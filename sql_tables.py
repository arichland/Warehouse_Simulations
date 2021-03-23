_author_ = 'arichland'
import pydict
import pymysql
import openpyxl




# Need to iterate through rows & insert into SQL
def excel_to_sql():
    # SQL Fields
    sql = pydict.localhost.get
    qry = pydict.sql_queries.get
    user = sql('user')
    pw = sql('password')
    host = sql('host')
    db = sql('database')

    # Excel Fields
    filename = "wms_dataset.xlsx"
    wb = openpyxl.load_workbook(filename=filename)
    wb.active = wb.sheetnames.index("orders")
    ws = wb.active
    for row in ws.rows:
        print(ws.rows[row])
        con = pymysql.connect(user=user, password=pw, host=host, database=db)
        with con.cursor() as cur:
            try:
                pass
                query = qry("orders")
                #cur.execute(query, )
            finally:
                con.commit()
            cur.close()
            con.close()






    sheets = wb.sheetnames
    col_range = {}
    data = {}
    count = 0

    # Get count of columns per sheet and save to dict
    for i in sheets:
        wb.active = wb.sheetnames.index(i)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            temp = {i: len(row)}
            col_range.update(temp)

    for k, v in col_range.items():
        wb.active = wb.sheetnames.index(k)
        ws = wb.active
        #print("sheet=", k)





        con = pymysql.connect(user=user, password=pw, host=host, database=db)
        with con.cursor() as cur:
            try:
                pass
                query = qry(k)
                #print(query)
            finally:
                    con.commit()
            cur.close()
            con.close()


def excel_dataset():
    # SQL Fields
    sql = pydict.localhost.get
    qry = pydict.sql_queries.get
    user = sql('user')
    pw = sql('password')
    host = sql('host')
    db = sql('database')

    # Excel Fields
    filename = "wms_dataset.xlsx"
    wb = openpyxl.load_workbook(filename=filename)
    wb.active = wb.sheetnames.index("orders")
    ws = wb.active
    for row in ws.rows:
        print(row)







    sheets = wb.sheetnames
    col_range = {}
    data = {}
    count = 0

    # Get count of columns per sheet and save to dict
    for i in sheets:
        wb.active = wb.sheetnames.index(i)
        ws = wb.active
        for row in ws.rows:
            print(row)

    for k, v in col_range.items():
        wb.active = wb.sheetnames.index(k)
        ws = wb.active
        #print("sheet=", k)





        con = pymysql.connect(user=user, password=pw, host=host, database=db)
        with con.cursor() as cur:
            try:
                pass
                query = qry(k)
                #print(query)
            finally:
                    con.commit()
            cur.close()
            con.close()





#excel_dataset()

def tbl_wms_master_data():
    sql = pydict.localhost.get
    user = sql('user')
    pw = sql('password')
    host = sql('host')
    db = sql('database')

    con = pymysql.connect(user=user, password=pw, host=host, database=db)
    with con.cursor() as cur:
        try:
            qry_create_table = """
            CREATE TABLE IF NOT EXISTS tbl_wms_master_data(
            id INT AUTO_INCREMENT PRIMARY KEY,
            reference_id INT,
            reference_name TEXT,
            storage_location TEXT)
            ENGINE=INNODB;"""
            cur.execute(qry_create_table)
        finally:
            con.commit()
    cur.close()
    con.close()

def tbl_wms_coordinates():
    sql = pydict.localhost.get
    user = sql('user')
    pw = sql('password')
    host = sql('host')
    db = sql('database')

    con = pymysql.connect(user=user, password=pw, host=host, database=db)
    with con.cursor() as cur:
        try:
            qry_create_table = """
            CREATE TABLE IF NOT EXISTS tbl_wms_coordinates(
            id INT AUTO_INCREMENT PRIMARY KEY,
            storage_location TEXT,
            x_coordinate INT,
            y_coordinate INT)
            ENGINE=INNODB;"""
            cur.execute(qry_create_table)
        finally:
            con.commit()
    cur.close()
    con.close()

def tbl_wms_orders():
    sql = pydict.localhost.get
    user = sql('user')
    pw = sql('password')
    host = sql('host')
    db = sql('database')

    con = pymysql.connect(user=user, password=pw, host=host, database=db)
    with con.cursor() as cur:
        try:
            qry_create_table = """
            CREATE TABLE IF NOT EXISTS tbl_wms_orders(
            id INT AUTO_INCREMENT PRIMARY KEY,
            timestamp DATETIME,
            reference_id INT,
            quantity INT,
            destination TEXT)
            ENGINE=INNODB;"""
            cur.execute(qry_create_table)
        finally:
            con.commit()
    cur.close()
    con.close()

def tables():
    tbl_wms_master_data()
    tbl_wms_coordinates()
    tbl_wms_orders()
tables()